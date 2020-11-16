#将筛选结果按会议分页存储至excel
import openpyxl
import os
import bibtexparser
from openpyxl.reader.excel import load_workbook
from openpyxl.workbook.workbook import Workbook
import pandas as pd
year_start = 2016
year_end = 2021
fields = [
    "AI",
    "NIS"    
]
excel_path = "筛选结果_可解释性.xlsx"
if not os.path.exists(excel_path) :
    wb = Workbook()
    wb.save(excel_path)

for field in fields:
    for year in range(year_start,year_end):
        current_dir = "filter_interpretability/{}/{}".format(field,year)
        for bibtex_file_name in os.listdir(current_dir):
            print(field,year,bibtex_file_name)
            with open(current_dir+'/'+bibtex_file_name) as bibtex_file:
                bib_database = bibtexparser.load(bibtex_file)
            if len(bib_database.entries) != 0:
                pf = pd.DataFrame(bib_database.entries)
                order=["title","title_cn","author","year","booktitle","url","pages","bibsource","biburl","timestamp"]
                pf = pf[order]
                pf.fillna(' ',inplace = True)

                wb = load_workbook(excel_path)
                all_sheets = wb.sheetnames
                with pd.ExcelWriter(excel_path,mode="a",engine="openpyxl") as writer:
                    writer.book = wb
                    writer.sheets = dict((ws.title, ws) for ws in wb.worksheets)
                    row = 0
                    sheet_name = bibtex_file_name.split("_")[0]
                    if sheet_name in all_sheets:
                        df = pd.DataFrame(pd.read_excel(excel_path,sheet_name=sheet_name))
                        row = df.shape[0]
                        print(row,sheet_name)
                        pf.to_excel(writer,encoding = 'utf-8',index = False,header = False,sheet_name=sheet_name,startrow=row+1)
                    else:
                        print(sheet_name)
                        pf.to_excel(writer,encoding = 'utf-8',index = False,sheet_name=sheet_name)
                writer.save()
                
wb = load_workbook(excel_path)
all_sheets = wb.sheetnames   
if "Sheet" in all_sheets:
    wb.remove(wb["Sheet"])
    wb.save(excel_path)        